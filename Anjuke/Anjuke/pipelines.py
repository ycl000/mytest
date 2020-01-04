# -*- coding: utf-8 -*-

# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: https://doc.scrapy.org/en/latest/topics/item-pipeline.html
from pprint import pprint
import pymysql
import datetime
from pymongo import MongoClient
from openpyxl import Workbook
# client = MongoClient(host="127.0.0.1",port=27017)
# collection = client["housedata"]["allinfo"]
wb = Workbook()
ws = wb.active
ws.title = "house_info"
ws["A1"] = "country"# 城市
ws["B1"] = "region"  # 区
ws["C1"] = "road_section"# 路段
ws["D1"] = "purpose" #房屋性质
ws["E1"] = "village" #小区名
# ws["F1"] = "图书照片链接"
ws["F1"] = "time_year"# 建筑时间
ws["G1"] = "price"  # 总价
ws["H1"] = "danjia"# 每平方米价格
ws["I1"] = "area"# 面积
ws["J1"] = "fangxing" #户型
ws["K1"] = "floor" #楼层
ws["L1"] = "describe" #描述
ws["M1"] = "house_count" #该区二手房数量
ws["N1"] = "orientation" #朝向
ws["O1"] = "people" #小区户数
ws["P1"] = "renovation" #装修类型
ws["Q1"] = "developers" #开发商
ws["R1"] = "house_label" #二手房标签




class AnjukePipeline(object):
    def process_item(self, item, spider):
        print("process_item")
        pprint(item)
        ws.append([item["country"], item["region"], item["road_section"], item["purpose"],
                   item["village"], item["time_year"], item["price"],
                   item["danjia"], item['area'], item['fangxing'],item['floor']
                   ,item['describe'],item['house_count'],item['orientation'],item['people'],
                   item['renovation'],item['developers'],item['house_label'],
                   ])
        wb.save("houseInfo1.xlsx")
        return item
